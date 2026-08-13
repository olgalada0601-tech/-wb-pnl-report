import numpy as np
import pandas as pd
from pathlib import Path
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path("/opt/wb-automation")
RAW_DIR = BASE_DIR / "data" / "raw"
REPORTS_DIR = BASE_DIR / "data" / "reports"

PRICES = {
    "17724581": 1290,
    "17724582": 1290,
    "18830011": 2490,
    "19902233": 890,
}

PRODUCTS = {
    "17724581": "Футболка оверсайз белая",
    "17724582": "Футболка оверсайз чёрная",
    "18830011": "Худи базовое бежевое",
    "19902233": "Кепка с логотипом",
}


def generate_demo_data() -> pd.DataFrame:
    np.random.seed(42)
    dates = pd.date_range("2026-07-01", periods=30, freq="D")
    rows = []
    for d in dates:
        for sku in PRODUCTS:
            price = PRICES[sku]
            qty = int(np.random.randint(3, 25))
            returns_qty = int(np.random.randint(0, 3))
            revenue = qty * price
            refunds = returns_qty * price
            commission = round(revenue * 0.15)
            logistics = (qty + returns_qty) * 62
            storage = round(revenue * 0.01)
            cost = round((qty - returns_qty) * price * 0.45)
            rows.append({
                "date": d.date(),
                "sku": sku,
                "product": PRODUCTS[sku],
                "qty": qty,
                "returns_qty": returns_qty,
                "revenue": revenue,
                "refunds": refunds,
                "commission": commission,
                "logistics": logistics,
                "storage": storage,
                "cost": cost,
            })
    return pd.DataFrame(rows)


def build_pnl(df: pd.DataFrame) -> dict:
    df = df.copy()
    df["net_profit"] = (
        df["revenue"] - df["refunds"] - df["commission"]
        - df["logistics"] - df["storage"] - df["cost"]
    )

    summary = pd.DataFrame([{
        "Выручка": df["revenue"].sum(),
        "Возвраты": df["refunds"].sum(),
        "Комиссия": df["commission"].sum(),
        "Логистика": df["logistics"].sum(),
        "Хранение": df["storage"].sum(),
        "Себестоимость": df["cost"].sum(),
        "Чистая прибыль": df["net_profit"].sum(),
        "Маржа %": round(df["net_profit"].sum() / df["revenue"].sum() * 100, 1),
    }])

    by_day = df.groupby("date", as_index=False)[["revenue", "net_profit"]].sum()
    by_day["date"] = by_day["date"].astype(str)
    by_day = by_day.rename(columns={
        "date": "Дата",
        "revenue": "Выручка",
        "net_profit": "Чистая прибыль",
    })

    by_sku = df.groupby(["sku", "product"], as_index=False)[["qty", "revenue", "net_profit"]].sum()
    by_sku["маржа %"] = round(by_sku["net_profit"] / by_sku["revenue"] * 100, 1)
    by_sku = by_sku.rename(columns={
        "sku": "Артикул",
        "product": "Товар",
        "qty": "Продажи, шт",
        "revenue": "Выручка",
        "net_profit": "Чистая прибыль",
    })

    return {"summary": summary, "by_day": by_day, "by_sku": by_sku}


def style_excel(path: Path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column_cells in ws.columns:
            values = [str(c.value) for c in column_cells if c.value is not None]
            width = min(max((len(v) for v in values), default=10) + 2, 40)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    wb.save(path)


def main():
    logger.info("Генерирую демо-данные WB...")
    df = generate_demo_data()
    raw_path = RAW_DIR / "wb_sales_demo.csv"
    df.to_csv(raw_path, index=False, encoding="utf-8-sig")
    logger.success(f"Сырые данные сохранены: {raw_path}")

    logger.info("Считаю P&L...")
    sheets = build_pnl(df)

    report_path = REPORTS_DIR / "pnl_demo.xlsx"
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        sheets["summary"].to_excel(writer, sheet_name="Сводка", index=False)
        sheets["by_day"].to_excel(writer, sheet_name="По дням", index=False)
        sheets["by_sku"].to_excel(writer, sheet_name="По артикулам", index=False)

    style_excel(report_path)
    logger.success(f"P&L отчёт создан: {report_path}")

    print()
    print(sheets["summary"].to_string(index=False))
    print()
    print(sheets["by_sku"].to_string(index=False))


if __name__ == "__main__":
    main()